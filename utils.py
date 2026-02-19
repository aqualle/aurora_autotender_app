import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import os
import shutil

from pathlib import Path
import sys


def normalize(text: str) -> str:
    """Приводим строку к нижнему регистру, убираем лишние пробелы и знаки."""
    if not isinstance(text, str):
        return ""
    text = text.lower()
    text = re.sub(r"[^a-zа-я0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def extract_products_from_excel(path: str):
    """Ищет лист с колонкой 'Наименование...' и возвращает товары."""
    all_sheets = pd.read_excel(path, header=None, sheet_name=None)
    found_df = None
    col_index = None
    start_row = None
    end_row = None

    for sheet_name, df in all_sheets.items():
        for i, row in df.iterrows():
            for j, val in row.items():
                if isinstance(val, str) and "наименование" in val.lower():
                    found_df = df
                    col_index = j
                    start_row = i + 1
                    break
            if found_df is not None:
                break
        if found_df is not None:
            break

    if found_df is None:
        raise ValueError("Не найден лист с колонкой 'Наименование...'")

    # ищем конец (строка 'Итого без НДС')
    for i, val in enumerate(found_df[col_index]):
        if isinstance(val, str) and "итого без ндс" in val.lower():
            end_row = i
            break

    if end_row is None:
        raise ValueError("Не найден конец таблицы ('Итого без НДС')")

    # собираем товары
    items = []
    for text in found_df.loc[start_row:end_row - 1, col_index]:
        if not isinstance(text, str):
            continue
        if text.lower().startswith("возможность поставки") or text.lower().startswith("валюта"):
            continue
        raw = text.strip()
        name = re.split(r"\n", raw)[0].strip()
        if name:
            items.append({"raw": raw, "name": name})

    return pd.DataFrame(items)

def save_results_into_excel(original_path: str, output_path: str, df: pd.DataFrame,
                           original_sheet_name="Original", prices_sheet_name="Prices"):
    """Сохраняет результат: лист Original + лист Prices"""
    original = pd.read_excel(original_path, header=None)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        original.to_excel(writer, sheet_name=original_sheet_name, index=False, header=False)
        df.to_excel(writer, sheet_name=prices_sheet_name, index=False)
    print(f"Результаты сохранены в {output_path}")

def get_merged_cell_value(ws, row, col):
    """Получает значение ячейки, даже если она объединена"""
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
                return ws.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value

def parse_price_value(price_str: str) -> float:
    """
    Извлекает числовое значение из строки цены.
    Обрабатывает форматы:
    - "39 419,17" → 39419.17
    - "42 512₽" → 42512.0
    - "42,705 ₽" → 42.705
    """
    if not price_str:
        return float('inf')
    
    price_str = str(price_str)
    
    # 1. Убираем неразрывные пробелы и обычные пробелы
    price_str = price_str.replace('\u00A0', '').replace(' ', '')
    
    # 2. Заменяем запятую на точку (русский формат: запятая = разделитель дробной части)
    price_str = price_str.replace(',', '.')
    
    # 3. Удаляем всё кроме цифр и точки
    clean = re.sub(r'[^0-9.]', '', price_str)
    
    try:
        return float(clean) if clean else float('inf')
    except:
        return float('inf')

def get_color_for_difference(difference: float, winner_price: float) -> str:
    """
    Возвращает цвет для раскраски РАЗНИЦЫ цен.
    
    difference = winner_price - our_price (положительное = мы дешевле, отрицательное = мы дороже)
    winner_price = цена победителя тендера
    
    Раскраска РАЗНИЦЫ:
    1) КРАСНЫЙ (FF0000): разница < 0 (наша цена > победителя)
    2) ЗЕЛЁНЫЙ (00B050): разница > 10% от winner_price (наша цена < победителя на >10%)
    3) ЖЁЛТЫЙ (FFFF00): разница 1-10% от winner_price (наша цена < победителя на 1-10%)
    """
    
    if difference < 0:
        # Наша цена больше (разница отрицательная) → КРАСНЫЙ
        return "00B050"
    
    # Наша цена меньше. Считаем на сколько процентов
    percentage_diff = (difference / winner_price * 100) if winner_price > 0 else 0
    
    if percentage_diff > 10:
        # Разница больше 10% → ЗЕЛЁНЫЙ (очень выгодно)
        return "FF0000"
    elif percentage_diff >= 1:
        # Разница 1-10% → ЖЁЛТЫЙ (выгодно)
        return "FFFF00"
    else:
        # Разница менее 1% → БЕЛЫЙ
        return "FFFFFF"

def find_yellow_field_row(ws, base_row: int, name_col: int) -> int:
    """Находит жёлтую ячейку для ссылки"""
    for offset in range(0, 13):
        check_row = base_row + offset
        if check_row > ws.max_row:
            return base_row + 3
        cell = ws.cell(check_row, name_col)
        if not isinstance(cell, MergedCell) and cell.fill and cell.fill.start_color:
            color = str(cell.fill.start_color.rgb) if hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color.index)
            if 'FFFF00' in color.upper() or 'FFEB9C' in color.upper():
                return check_row
    return base_row + 3

def find_or_create_marketplace_column(ws, header_row: int, name_col: int, column_name: str) -> int:
    """Находит СУЩЕСТВУЮЩУЮ колонку маркетплейса или создаёт новую"""
    # ИЩЕМ СУЩЕСТВУЮЩУЮ колонку маркетплейса
    for col_idx in range(name_col + 1, ws.max_column + 2):
        h = get_merged_cell_value(ws, header_row, col_idx)
        if h and isinstance(h, str) and column_name in h:
            print(f"Колонка '{column_name}' найдена: {get_column_letter(col_idx)}")
            return col_idx

    # СОЗДАЁМ новую колонку ПОСЛЕ всех маркетплейсов и участников
    max_col = name_col + 1
    for col_idx in range(name_col + 1, ws.max_column + 10):
        h = get_merged_cell_value(ws, header_row, col_idx)
        if h and isinstance(h, str) and h.strip():
            max_col = col_idx + 1

    new_col = max_col

    # Заголовок
    header_cell = ws.cell(header_row, new_col)
    header_cell.value = column_name
    header_cell.font = Font(bold=True)
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    print(f"📍 Колонка '{column_name}' создана: {get_column_letter(new_col)}")
    return new_col

def find_or_create_difference_column(ws, header_row: int, marketplace_col: int, column_name: str) -> int:
    """Создаёт колонку для разницы цен (ПОСЛЕ колонки маркетплейса)"""
    new_col = marketplace_col + 1

    # Заголовок
    header_cell = ws.cell(header_row, new_col)
    header_cell.value = column_name
    header_cell.font = Font(bold=True)
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    print(f"📍 Колонка '{column_name}' создана: {get_column_letter(new_col)}")
    return new_col

def save_results_into_tender_format(original_path: str, output_path: str,
                                   df: pd.DataFrame, target_sheet_name: str = None,
                                   column_name: str = "Яндекс Маркет"):
    """
    Сохраняет результаты парсинга в тендерную таблицу с форматированием.
    Добавляет ДВЕ колонки:
    1. Цена (БЕЗ раскраски - просто число)
    2. Разница цен (С РАСКРАСКОЙ)
    
    Раскраска РАЗНИЦЫ:
    - КРАСНЫЙ: наша цена > победителя (разница < 0)
    - ЗЕЛЁНЫЙ: наша цена < победителя на >10% (разница > 10%)
    - ЖЁЛТЫЙ: наша цена < победителя на 1-10% (разница 1-10%)
    """
    print(f"📋 Создаю колонки для '{column_name}' в тендерной таблице...")

    is_yandex = "яндекс" in column_name.lower()
    is_ozon = "ozon" in column_name.lower()

    if is_yandex:
        print("🔗 Режим: ГИПЕРССЫЛКА для Яндекс.Маркет")
    elif is_ozon:
        print("🔗 Режим: ГИПЕРССЫЛКА для Ozon")

    try:
        if not os.path.exists(output_path):
            shutil.copy2(original_path, output_path)
            print("✅ Файл скопирован")

        wb = load_workbook(output_path)
        ws = wb[target_sheet_name] if target_sheet_name and target_sheet_name in wb.sheetnames else wb.active

        # Находим колонку "Наименование"
        name_col = None
        name_start_row = None
        number_col = None

        for row_idx in range(1, 21):
            for col_idx in range(1, 11):
                val = get_merged_cell_value(ws, row_idx, col_idx)
                if val and isinstance(val, str):
                    if 'наименование' in val.lower():
                        name_col = col_idx
                        name_start_row = row_idx + 1
                    if '№' in val:
                        number_col = col_idx

        if not name_col:
            raise ValueError("Не найдена колонка Наименование")

        header_row = name_start_row - 1

        # НАХОДИМ или СОЗДАЁМ колонку маркетплейса
        marketplace_col = find_or_create_marketplace_column(ws, header_row, name_col, column_name)

        # СОЗДАЁМ колонку разницы цен
        difference_col = find_or_create_difference_column(ws, header_row, marketplace_col, f"Разница {column_name}")

        # Находим участников для сравнения цен
        participants = []
        for col_idx in range(name_col + 1, marketplace_col):
            h = get_merged_cell_value(ws, header_row, col_idx)
            if h and isinstance(h, str) and h.strip() and column_name not in h:
                participants.append({'col': col_idx, 'name': h.strip()})

        print(f"📊 Найдено участников: {len(participants)}")

        # Заполняем данные
        filled_count = 0
        link_count = 0

        for idx, (_, item) in enumerate(df.iterrows()):
            position = idx + 1
            
            # ИЩЕМ РЯДОК ТОВАРА ПО НОМЕРУ ПОЗИЦИИ
            base_row = None
            for row_idx in range(name_start_row, ws.max_row):
                pos_cell = get_merged_cell_value(ws, row_idx, number_col if number_col else 1)
                if pos_cell and str(pos_cell).strip() == str(position):
                    base_row = row_idx
                    break
            
            if not base_row:
                print(f"⚠️ Не найдена строка товара #{position}")
                continue

            price = item.get('цена', '')
            price_vat = item.get('цена для юрлиц', '')
            link = item.get('ссылка', '')

            # Находим победителя для сравнения (ищем "1 место" в рядке товара)
            winner_col = None
            min_price_without = float('inf')
            min_price_with = float('inf')

            for p in participants:
                rank = get_merged_cell_value(ws, base_row, p['col'])
                if rank and '1' in str(rank) and 'место' in str(rank).lower():
                    winner_col = p['col']
                    break

            if winner_col:
                # БЕРЁМ ПОБЕДИТЕЛЯ ИЗ ТЕХ ЖЕ РЯДОВ ГДЕ ИЩЕМ НАШУ ЦЕНУ
                # base_row + 1 = рядок "Цена без НДС"
                # base_row + 2 = рядок "Цена с НДС"
                
                p1 = get_merged_cell_value(ws, base_row + 1, winner_col)
                if p1:
                    min_price_without = parse_price_value(str(p1))

                p2 = get_merged_cell_value(ws, base_row + 2, winner_col)
                if p2:
                    min_price_with = parse_price_value(str(p2))

            # ==================== ЦЕНА БЕЗ НДС ====================
            if price:
                c = ws.cell(base_row + 1, marketplace_col)
                if not isinstance(c, MergedCell):
                    price_num = parse_price_value(price)
                    c.value = price_num
                    c.alignment = Alignment(horizontal='right')
                    filled_count += 1

                diff_cell = ws.cell(base_row + 1, difference_col)
                if not isinstance(diff_cell, MergedCell):
                    price_num = parse_price_value(price)
                    
                    if min_price_without != float('inf'):
                        difference = min_price_without - price_num
                        diff_cell.value = int(difference)
                        diff_cell.alignment = Alignment(horizontal='right')
                        
                        color = get_color_for_difference(difference, min_price_without)
                        diff_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # ==================== ЦЕНА С НДС ====================
            if price_vat:
                c = ws.cell(base_row + 2, marketplace_col)
                if not isinstance(c, MergedCell):
                    price_vat_num = parse_price_value(price_vat)
                    c.value = price_vat_num
                    c.alignment = Alignment(horizontal='right')

                diff_cell = ws.cell(base_row + 2, difference_col)
                if not isinstance(diff_cell, MergedCell):
                    price_vat_num = parse_price_value(price_vat)
                    
                    if min_price_with != float('inf'):
                        difference = min_price_with - price_vat_num
                        diff_cell.value = int(difference)
                        diff_cell.alignment = Alignment(horizontal='right')
                        
                        color = get_color_for_difference(difference, min_price_with)
                        diff_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # ==================== ССЫЛКА ====================
            if link:
                yellow_row = find_yellow_field_row(ws, base_row, name_col)
                link_cell = ws.cell(yellow_row, marketplace_col)

                if not isinstance(link_cell, MergedCell):
                    if is_yandex:
                        link_cell.value = "Ссылка"
                        link_cell.hyperlink = link
                        link_cell.font = Font(color="0563C1", underline="single", size=9)
                        link_cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif is_ozon:
                        link_cell.value = "Ссылка"
                        link_cell.hyperlink = link
                        link_cell.font = Font(color="0563C1", underline="single", size=9)
                        link_cell.alignment = Alignment(horizontal='center', vertical='center')

                    link_count += 1

        # Границы для обеих колонок
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        for row_idx in range(header_row, ws.max_row + 1):
            for col_idx in [marketplace_col, difference_col]:
                c = ws.cell(row_idx, col_idx)
                if not isinstance(c, MergedCell):
                    c.border = border

        wb.save(output_path)

        print(f"✅ Заполнено: {filled_count} товаров")
        print(f"🔗 Сохранено ссылок: {link_count}")
        print(f"💾 Сохранено: {output_path}")

        return True

    except Exception as e:
        print(f"❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        return False

from pathlib import Path
import sys


def get_app_dir() -> Path:
    """
    Возвращает папку приложения:
    - при запуске .exe (PyInstaller)
    - при запуске из .py
    """
    if getattr(sys, 'frozen', False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent


def get_browser_paths():
    import sys
    from pathlib import Path

    # 🔥 базовая директория
    if getattr(sys, 'frozen', False):
        base_dir = Path(sys._MEIPASS)   # файлы, вшитые в exe
    else:
        base_dir = Path(__file__).parent

    paths = {
        "edge": {
#            "binary": Path(
#                r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
#            ),
             "binary": base_dir / "browser" / "edge" / "msedge.exe",
            "driver": base_dir / "browserdriver" / "msedgedriver.exe",
        },
        "chrome": {
            "binary": Path(
                r"C:\Program Files\Google\Chrome\Application\chrome.exe"
            ),
            "driver": base_dir / "browserdriver" / "chromedriver.exe",
        },
    }

    return paths